from openpyxl import load_workbook

workbook = load_workbook(filename = "Agenda_Export.xlsx")
sheet = workbook.active

# Filter list to remove session blocks, keep only talks/posters    
contributions = [row for row in sheet.iter_rows(min_row=2, values_only=True)
                 if int(row[8])>0]

f = open("Contributions.md","w")
f.writelines("""---
pagetitle: List of all contributions
---

# List of all contributions

## Plenaries

""")

day = {"29" : "Monday, July 29",
       "30" : "Tuesday, July 30",
       "31" : "Wednesday, July 31",
       "01" : "Thursday, August 1",
       "02" : "Friday, August 2"}

def write_name(c):
    i = 20
    while not int(c[i+10]):
        i += 13    
    name = c[i] + " " + c[i+1]
    title = c[13].lstrip()
    f.writelines(f'''<form class="form-horizontal" name="schedule_search"
    method="POST" action="https://express.converia.de/frontend/index.php">
    <input type="hidden" name="page_id" value="37679">
    <input type="hidden" name="v" value="List">
    <input type="hidden" name="day" value="all">
    <input type="hidden" name="do" value="13">
    <input type="hidden" name="q" value="{title}">
    <input type="submit" value="{name}"> </form>\n''')
    f.write(c[13].lstrip() + "\n")
    if c[23] is not None:
        f.write ("  ~ " + c[23].lstrip() + '\n')
    
for c in contributions:
    if c[12] == "1. Plenary Speakers":
        write_name(c)
        f.write("  ~ " + day[c[3][0:2]] + ", " + c[1] + "-" + \
                c[2] + ", " + c[10] + "\n")
        f.write("\n")

sessions = [c[12] for c in contributions]
sessions = list(set(sessions))
sessions.sort()
minisymposia = [s for s in sessions if s[0:4] == "Mini"]
contributed = [s for s in sessions if s[0:4] == "Sess"]

f.write("## Minisymposia")
f.write("\n")

for t in minisymposia:
    f.write ("#### " + t + "\n")
    for c in contributions:
        if c[12] == t:
            write_name(c)
            f.write("  ~ " + day[c[3][0:2]] + ", " + c[1] + "-" + \
                    c[2] + ", " + c[10] + "\n")
#            f.write("  ~ &nbsp;\n")
            f.write("\n")

        
f.write("## Contributed Sessions")
f.write("\n")

for t in contributed:
    f.write ("#### " + t + "\n")
    for c in contributions:
        if c[12] == t:
            write_name(c)
            f.write("  ~ " + day[c[3][0:2]] + ", " + c[1] + "-" + \
                    c[2] + ", " + c[10] + "\n")
#            f.write("  ~ &nbsp;\n")
            f.write("\n")

f.write("## Posters")
f.write("\n")

for c in contributions:
    if c[12][0:4] == "Post":
        write_name(c)
        f.write("  ~ " + day[c[3][0:2]] + ", " + \
                "16:00-18:30, Campus Center (IRC) Foyer\n")
#        f.write("  ~ &nbsp;\n")
        f.write("\n")

f.close()
